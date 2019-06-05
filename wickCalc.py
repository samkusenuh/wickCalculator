#!/usr/bin/env python
# coding: utf-8

# In[48]:


import openpyxl
import numpy as np
import math
from matplotlib import pyplot as plt
from matplotlib.patches import Polygon
import os
import csv


# In[49]:


print('Welcome to Sams Wick Calculator')
print('Lets try to find your chart')
print('Your current working directory is ' + str(os.getcwd()))
print('if your chart is not in this current directory, please copy the full path and paste it below')
chartDir = input(r"Please paste the full path of the FOLDER like so: C:\Users\...\...")

os.chdir(chartDir)

print('your current directory has become ' + str(os.getcwd()))


# In[50]:


print ('now paste the csv file here')
csvFile = input('paste it including the .csv extension:  ')


# In[51]:


wb = openpyxl.Workbook()
ws = wb.active

with open(csvFile) as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)
print('We are now making an xlsx file of your csv file...')

newfileName = csvFile[0:-4] + '.xlsx'

wb.save(newfileName)
print('conversion complete!')




wb = openpyxl.load_workbook(newfileName)
ws = wb['Sheet']


# In[53]:


print('retrieving all the wicks...')
wickList = []

for row in range(2, ws.max_row + 1):

    openPrice = ws['B' + str(row)].value
    closePrice = ws['E' + str(row)].value
    highPrice = ws['C' + str(row)].value
    lowPrice = ws['D' + str(row)].value

    if float(openPrice) < float(closePrice): #bullish candle bar

        upperWickLength = float(highPrice) - float(closePrice)
        upperWickLength = float("{0:.6f}".format(upperWickLength))
        wickList.append(upperWickLength)
        lowerWickLength = float(openPrice) - float(lowPrice)
        lowerWickLength = float("{0:.6f}".format(lowerWickLength))
        wickList.append(lowerWickLength)


    elif float(openPrice) > float(closePrice): #bearish candle bar

        upperWickLength = float(highPrice) - float(openPrice)
        upperWickLength = float("{0:.6f}".format(upperWickLength))
        wickList.append(upperWickLength)
        lowerWickLength = float(lowPrice) - float(lowPrice)
        lowerWickLength = float("{0:.6f}".format(lowerWickLength))
        wickList.append(lowerWickLength)

    else:
        upperWickLength = float(highPrice) - float(openPrice)
        upperWickLength = float("{0:.6f}".format(upperWickLength))
        wickList.append(upperWickLength)
        lowerWickLength = float(lowPrice) - float(lowPrice)
        lowerWickLength = float("{0:.6f}".format(lowerWickLength))
        wickList.append(lowerWickLength)


print('wick extraction complete')
print(str(max(wickList)) + '  ' + str(min(wickList)) + '  ' + str(np.mean(wickList)) + '  ' + str(np.std(wickList))   )

print(len(wickList))

cleanWickList = [x for x in wickList if x != 0.0]

print(len(cleanWickList))











# In[54]:


bins = np.linspace(math.ceil(min(cleanWickList)),
                   math.floor(max(cleanWickList)),
                   200) # fixed number of bins

plt.xlim([min(cleanWickList), max(cleanWickList)])

plt.hist(cleanWickList, bins = 200)
plt.title('Random Gaussian data (fixed number of bins)')
plt.xlabel('variable X (20 evenly spaced bins)')
plt.ylabel('count')

plt.show()


# In[55]:


# basic plot

plt.boxplot(cleanWickList)


# In[56]:


ninetyNinePercent = np.percentile(cleanWickList, 99.9)  # Q1

hunderedPercent = np.percentile(cleanWickList, 100)


# In[57]:


print("the largest wicks (0.1%) are between " + str(float("{0:.4f}".format(ninetyNinePercent))) + ' and ' + str(hunderedPercent) +  ' long ')




# In[ ]:





# In[ ]:
